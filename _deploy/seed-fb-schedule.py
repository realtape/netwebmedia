#!/usr/bin/env python3
"""
Seed all PENDING posts from NetWebMedia_Posts_Tracker.xlsx into Facebook Page
scheduled posts via the fb_publish handler.

Reads the tracker, builds payloads, POSTs to:
    https://netwebmedia.com/crm-vanilla/api/?r=fb_publish&action=schedule

Assumes:
  - Assets have already been deployed under /assets/social/campaign/ (run after FTPS deploy).
  - FB_PAGE_ID + FB_PAGE_TOKEN secrets are wired (handler returns 503 otherwise).
  - MIGRATE_TOKEN env var is set to the live value (matches config.local.php).

Usage:
    MIGRATE_TOKEN=... python3 _deploy/seed-fb-schedule.py [--dry-run] [--api-base URL] [--asset-base URL]

Defaults:
    --api-base    https://netwebmedia.com/crm-vanilla/api
    --asset-base  https://netwebmedia.com/assets/social/campaign
    --tracker     C:/Users/Usuario/Documents/Claude/Projects/Social Media Campaign/NetWebMedia_Posts_Tracker.xlsx

Time zone:
    The tracker's "Schedule Time" column is interpreted as America/Santiago (Chile)
    local time. May = UTC-3 (no DST). Override with --tz if needed.
"""
import argparse
import json
import os
import sys
import urllib.request
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

# Chile (America/Santiago) is UTC-3 in May (winter, no DST observed since 2022).
# Hardcoded offset avoids tzdata dependency on Windows. Override via --tz-offset.
DEFAULT_TZ_OFFSET_HOURS = -3

DEFAULT_TRACKER = r"C:\Users\Usuario\Documents\Claude\Projects\Social Media Campaign\NetWebMedia_Posts_Tracker.xlsx"
DEFAULT_API_BASE = "https://netwebmedia.com/crm-vanilla/api"
DEFAULT_ASSET_BASE = "https://netwebmedia.com/assets/social/campaign"

# Carousel themes → folder name + slide count.
# Slide files are named "<folder>_slide_<N>.png" (per local inventory).
CAROUSEL_FOLDERS = {
    "AEO Visibility Checklist": {
        "EN": "aeo_en",
        "ES": "aeo_es",
    },
    "CMO Growth Engagement Breakdown": {
        "EN": "cmo_growth_en",
        "ES": "cmo_growth_es",
    },
    "CMO Scale Engagement Structure": {
        "EN": "cmo_scale_en",
        "ES": "cmo_scale_es",
    },
}
CAROUSEL_SLIDES = 7

# Reel filename map per post (from tracker rows + local _final.mp4 filenames).
REEL_FILENAMES = {
    3: "reel_2_growth_en_final.mp4",
    4: "reel_2_growth_es_final.mp4",
    5: "reel_3_scale_en_final.mp4",
    6: "reel_3_scale_es_final.mp4",
}


def parse_xlsx(path: str) -> list[dict]:
    """Read 'All Posts' sheet, return list of row dicts (skipping header)."""
    wb = load_workbook(path, data_only=True)
    ws = wb["All Posts"]
    headers = [str(c.value or "").strip() for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = {h: (str(v).strip() if v is not None else "") for h, v in zip(headers, r)}
        rows.append(d)
    return rows


def parse_schedule_dt(date_str: str, time_str: str, tz_offset_hours: int) -> int:
    """
    Parse "2026-05-08 (Fri)" + "10:00 AM" at fixed UTC offset, return UNIX timestamp.
    """
    date_part = date_str.split("(")[0].strip()
    combined = f"{date_part} {time_str}"
    dt_naive = datetime.strptime(combined, "%Y-%m-%d %I:%M %p")
    tz = timezone(timedelta(hours=tz_offset_hours))
    return int(dt_naive.replace(tzinfo=tz).timestamp())


def build_post_payload(row: dict, asset_base: str, tz_offset_hours: int) -> dict | None:
    """Convert one tracker row into a schedule API payload entry."""
    status = row.get("Status", "").upper()
    if status != "PENDING":
        return None

    try:
        post_num = int(row["Post #"])
    except (ValueError, KeyError):
        return None

    fmt_label = row.get("Format", "").lower()
    is_carousel = "carousel" in fmt_label
    is_reel = "reel" in fmt_label

    caption = row.get("FB-tuned Caption (alternate, fewer hashtags)", "").strip()
    if not caption:
        caption = row.get("Caption (paste-ready, includes hashtags)", "").strip()
    if not caption:
        return {"_error": f"post {post_num}: no caption found"}

    sched_unix = parse_schedule_dt(
        row.get("Schedule Date", ""),
        row.get("Schedule Time (Chile = ET in May)", ""),
        tz_offset_hours,
    )

    payload = {
        "post_number": post_num,
        "scheduled_at_unix": sched_unix,
        "caption": caption,
    }

    if is_reel:
        fname = REEL_FILENAMES.get(post_num)
        if not fname:
            return {"_error": f"post {post_num}: no reel filename mapped"}
        payload["format"] = "video"
        payload["video_url"] = f"{asset_base.rstrip('/')}/{fname}"
        return payload

    if is_carousel:
        theme = row.get("Theme", "").strip()
        lang = row.get("Language", "").strip().upper()
        folder = CAROUSEL_FOLDERS.get(theme, {}).get(lang)
        if not folder:
            return {"_error": f"post {post_num}: no carousel folder mapped for theme={theme!r} lang={lang!r}"}
        payload["format"] = "carousel"
        payload["image_urls"] = [
            f"{asset_base.rstrip('/')}/carousels/{folder}/{folder}_slide_{i}.png"
            for i in range(1, CAROUSEL_SLIDES + 1)
        ]
        return payload

    return {"_error": f"post {post_num}: unknown format {fmt_label!r}"}


def post_schedule(api_base: str, token: str, posts: list[dict], dry_run: bool) -> dict:
    url = f"{api_base.rstrip('/')}/?r=fb_publish&action=schedule&token={token}"
    body = json.dumps({"posts": posts, "dry_run": dry_run}).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=body,
        method="POST",
        headers={
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0 (NWM seed-fb-schedule.py)",
            "Origin": "https://netwebmedia.com",
            "Referer": "https://netwebmedia.com/crm-vanilla/",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=600) as resp:
            return {"http": resp.status, "body": json.loads(resp.read().decode("utf-8"))}
    except urllib.error.HTTPError as e:
        return {"http": e.code, "body": e.read().decode("utf-8", errors="replace")}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tracker", default=DEFAULT_TRACKER)
    ap.add_argument("--api-base", default=DEFAULT_API_BASE)
    ap.add_argument("--asset-base", default=DEFAULT_ASSET_BASE)
    ap.add_argument("--tz-offset", type=int, default=DEFAULT_TZ_OFFSET_HOURS, help="UTC offset in hours (default -3 = Chile in May).")
    ap.add_argument("--dry-run", action="store_true", help="Send dry_run=true; logs to fb_publish_log without calling FB API.")
    ap.add_argument("--print-only", action="store_true", help="Print payloads, do not POST.")
    args = ap.parse_args()

    if not Path(args.tracker).exists():
        print(f"Tracker not found: {args.tracker}", file=sys.stderr)
        sys.exit(1)

    token = os.environ.get("MIGRATE_TOKEN", "").strip()
    if not args.print_only and not token:
        print("MIGRATE_TOKEN env var required (matches config.local.php).", file=sys.stderr)
        sys.exit(2)

    rows = parse_xlsx(args.tracker)
    payloads = []
    errors = []
    for row in rows:
        p = build_post_payload(row, args.asset_base, args.tz_offset)
        if p is None:
            continue
        if "_error" in p:
            errors.append(p["_error"])
            continue
        payloads.append(p)

    print(f"Parsed {len(payloads)} PENDING posts from tracker.")
    if errors:
        print(f"Skipped {len(errors)} with errors:")
        for e in errors:
            print(f"  - {e}")

    if args.print_only:
        print(json.dumps(payloads, indent=2))
        return

    print(f"\nPOSTing to {args.api_base} (dry_run={args.dry_run})…")
    result = post_schedule(args.api_base, token, payloads, args.dry_run)
    print(f"HTTP {result['http']}")
    print(json.dumps(result["body"], indent=2) if isinstance(result["body"], dict) else result["body"])


if __name__ == "__main__":
    main()
