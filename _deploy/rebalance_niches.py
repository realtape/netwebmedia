#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Rebalance 680 prospects across 8 niches (target: ~85 per niche).

Strategy:
- Keep Tourism/Restaurants/Health/Beauty as primary classification
- Allocate SMB, Law Firms, Real Estate, Local Specialist via keyword detection
- Respect regional niche affinity from niches_regions.py
- Use round-robin fallback for unclassified

Usage:
    python rebalance_niches.py --dry-run
    python rebalance_niches.py
"""
import json
import os
import sys
from collections import defaultdict

sys.path.insert(0, os.path.dirname(__file__))
from niches_regions import NICHES, NICHE_ORDER, REGIONS, CITY_TO_REGION

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(HERE, "NetWebMedia_Chile_680_Businesses_Website_Social_Audit.xlsx")
CRM_JSON_PATH = os.path.join(HERE, "crm_import.json")

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed")
    sys.exit(1)

def classify_niche_smart(name, vertical, city):
    """Classify with keyword + region affinity."""
    name_lower = name.lower()
    vertical_lower = vertical.lower() if vertical else ""

    # Hard rules: keyword detection (highest confidence)
    if any(kw in name_lower for kw in ["abogad", "law", "legal", "notari", "derecho"]):
        return "law_firms"
    if any(kw in name_lower for kw in ["inmobil", "propied", "corretaje", "real estate"]):
        return "real_estate"
    if any(kw in name_lower for kw in ["plumb", "fontaner", "electrician", "electricista", "hvac", "maestro", "pest"]):
        return "local_specialist"
    if any(kw in name_lower for kw in ["tienda", "store", "boutique", "shop", "comercio"]):
        return "smb"

    # Vertical-based fallback
    if "tourism" in vertical_lower or "turismo" in vertical_lower:
        return "tourism"
    if "restaurant" in vertical_lower or "gastro" in vertical_lower:
        return "restaurants"
    if "health" in vertical_lower or "salud" in vertical_lower or "medic" in vertical_lower:
        return "health"
    if "beauty" in vertical_lower or "belleza" in vertical_lower or "estetic" in vertical_lower:
        return "beauty"

    # Regional affinity: return first niche in region's list (respects REGIONS mapping)
    try:
        region_key = CITY_TO_REGION.get(city)
        if region_key and region_key in REGIONS:
            region_niches = REGIONS[region_key].get("niches", [])
            if region_niches:
                # Return first niche in region's preference list
                return region_niches[0]
    except:
        pass

    # Fallback to SMB (catch-all)
    return "smb"

def load_xlsx_and_classify():
    """Load XLSX and classify all records."""
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    records = []
    for row_idx in range(2, ws.max_row + 1):
        if not ws[row_idx][2].value:
            continue

        city = str(ws[row_idx][0].value or "").lower().replace(" ", "-")
        vertical = str(ws[row_idx][1].value or "")
        name = str(ws[row_idx][2].value or "")
        website = str(ws[row_idx][3].value or "")
        email = str(ws[row_idx][6].value or "") if ws[row_idx][6] else None

        niche_key = classify_niche_smart(name, vertical, city)

        records.append({
            "name": name,
            "email": email,
            "phone": str(ws[row_idx][7].value or "") if ws[row_idx][7] else "",
            "company": name,
            "role": f"{NICHES[niche_key]['name']} — {city.replace('-', ' ').title()}",
            "status": "lead",
            "value": 0,
            "niche_key": niche_key,
            "niche": NICHES[niche_key]["name"],
            "city": city,
            "website": website,
            "notes": json.dumps({
                "city": city,
                "niche": NICHES[niche_key]["name"],
                "website": website,
                "page": f"companies/{city}/{name.lower().replace(' ', '-')[:60]}.html",
            }, ensure_ascii=False),
        })

    wb.close()
    return records

def rebalance_to_target(records, target_per_niche=85):
    """Rebalance to approximately equal distribution across 8 niches."""
    # Current count
    by_niche = defaultdict(list)
    for r in records:
        by_niche[r["niche_key"]].append(r)

    print(f"Initial distribution:")
    for niche in NICHE_ORDER:
        count = len(by_niche.get(niche, []))
        print(f"  {niche:20s}: {count:3d}")

    # Target distribution: ~85 per niche (680 / 8)
    # Strategy: For niches with surplus, move to niches with deficit
    target = target_per_niche
    final_allocation = {n: [] for n in NICHE_ORDER}

    # First pass: allocate up to target from each niche's primary assignment
    remaining = []
    for niche in NICHE_ORDER:
        allocated = by_niche.get(niche, [])[:target]
        final_allocation[niche] = allocated
        remaining.extend(by_niche.get(niche, [])[target:])

    # Second pass: distribute remaining records to fill deficits by reassigning niche_key
    for niche in NICHE_ORDER:
        needed = target - len(final_allocation[niche])
        if needed > 0 and remaining:
            additions = remaining[:needed]
            # Reassign the niche_key and niche fields for these records
            for rec in additions:
                rec["niche_key"] = niche
                rec["niche"] = NICHES[niche]["name"]
                # Update role and notes
                rec["role"] = f"{NICHES[niche]['name']} — {rec.get('city', 'unknown').replace('-', ' ').title()}"
                if isinstance(rec.get("notes"), str):
                    import json
                    notes_obj = json.loads(rec["notes"])
                else:
                    notes_obj = rec.get("notes", {})
                notes_obj["niche"] = NICHES[niche]["name"]
                rec["notes"] = json.dumps(notes_obj, ensure_ascii=False)
            final_allocation[niche].extend(additions)
            remaining = remaining[needed:]

    # Third pass: distribute any stragglers evenly
    idx = 0
    while remaining:
        niche = NICHE_ORDER[idx % len(NICHE_ORDER)]
        rec = remaining.pop(0)
        rec["niche_key"] = niche
        rec["niche"] = NICHES[niche]["name"]
        rec["role"] = f"{NICHES[niche]['name']} — {rec.get('city', 'unknown').replace('-', ' ').title()}"
        if isinstance(rec.get("notes"), str):
            import json
            notes_obj = json.loads(rec["notes"])
        else:
            notes_obj = rec.get("notes", {})
        notes_obj["niche"] = NICHES[niche]["name"]
        rec["notes"] = json.dumps(notes_obj, ensure_ascii=False)
        final_allocation[niche].append(rec)
        idx += 1

    print(f"\nFinal balanced distribution:")
    total = 0
    for niche in NICHE_ORDER:
        count = len(final_allocation[niche])
        total += count
        print(f"  {niche:20s}: {count:3d}")
    print(f"  {'TOTAL':20s}: {total:3d}")

    # Flatten
    rebalanced = []
    for niche in NICHE_ORDER:
        rebalanced.extend(final_allocation[niche])

    return rebalanced

def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    print("Loading and classifying XLSX...")
    records = load_xlsx_and_classify()
    print(f"Loaded {len(records)} records")

    print("\nRebalancing across niches...")
    rebalanced = rebalance_to_target(records, target_per_niche=680//8)

    if args.dry_run:
        print("\n[DRY RUN] No files written")
        return

    # Write CRM JSON
    print(f"\nWriting {CRM_JSON_PATH}...")
    with open(CRM_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(rebalanced, f, indent=2, ensure_ascii=False)

    print(f"✓ Rebalanced crm_import.json written")

if __name__ == "__main__":
    main()
