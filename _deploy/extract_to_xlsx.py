#!/usr/bin/env python3
"""
Extract all business data from 17 city prospect HTML reports
and create an XLSX spreadsheet for Google Sheets import.
"""

import re
import os
from html.parser import HTMLParser
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DEPLOY_DIR = os.path.dirname(os.path.abspath(__file__))

CITIES = [
    "santiago", "valparaiso", "concepcion", "antofagasta", "iquique",
    "la-serena", "rancagua", "temuco", "puerto-montt", "arica",
    "copiapo", "talca", "chillan", "osorno", "valdivia",
    "punta-arenas", "coyhaique"
]

CITY_DISPLAY = {
    "santiago": "Santiago",
    "valparaiso": "Valparaíso",
    "concepcion": "Concepción",
    "antofagasta": "Antofagasta",
    "iquique": "Iquique",
    "la-serena": "La Serena",
    "rancagua": "Rancagua",
    "temuco": "Temuco",
    "puerto-montt": "Puerto Montt",
    "arica": "Arica",
    "copiapo": "Copiapó",
    "talca": "Talca",
    "chillan": "Chillán",
    "osorno": "Osorno",
    "valdivia": "Valdivia",
    "punta-arenas": "Punta Arenas",
    "coyhaique": "Coyhaique"
}

VERTICALS = ["Tourism", "Restaurants", "Health", "Beauty"]

def clean_html(text):
    """Remove HTML tags and clean whitespace."""
    text = re.sub(r'<[^>]+>', '', text)
    text = re.sub(r'&bull;', '•', text)
    text = re.sub(r'&amp;', '&', text)
    text = re.sub(r'&lt;', '<', text)
    text = re.sub(r'&gt;', '>', text)
    text = re.sub(r'&rarr;', '→', text)
    text = re.sub(r'&mdash;', '—', text)
    text = re.sub(r'&ndash;', '–', text)
    text = re.sub(r'&#8203;', '', text)
    text = re.sub(r'&nbsp;', ' ', text)
    text = re.sub(r'&#x27;', "'", text)
    text = re.sub(r'&#39;', "'", text)
    text = re.sub(r'&quot;', '"', text)
    text = re.sub(r'★', '★', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def extract_businesses_from_html(html_content, city_slug):
    """Extract business data from a prospect report HTML file."""
    businesses = []

    # Try biz-card format first, then prospect-card
    card_format = 'biz-card'
    cards = re.split(r'<div class="biz-card">', html_content)
    if len(cards) < 2:
        cards = re.split(r'<div class="prospect-card">', html_content)
        card_format = 'prospect-card'
    if len(cards) < 2:
        cards = re.split(r'class="biz-card"', html_content)
    if len(cards) < 2:
        cards = re.split(r'class="prospect-card"', html_content)
        card_format = 'prospect-card'

    # Skip the first split (before first card)
    cards = cards[1:]

    # Determine which section (vertical) each card belongs to
    # We need to find section headers in the original HTML
    section_markers = []
    for v in VERTICALS:
        # Look for section headers like "Tourism" "Restaurants" etc.
        patterns = [
            rf'🏨.*?Tourism',
            rf'🍽.*?Restaurant',
            rf'🏥.*?Health',
            rf'💇.*?Beauty',
            rf'<h2[^>]*>.*?Tourism',
            rf'<h2[^>]*>.*?Restaurant',
            rf'<h2[^>]*>.*?Health',
            rf'<h2[^>]*>.*?Beauty',
        ]

    # Find positions of section headers in original HTML
    tourism_pos = -1
    restaurant_pos = -1
    health_pos = -1
    beauty_pos = -1

    for pattern in [r'Tourism.*?(?:10 businesses|10 Businesses|Prospects)',
                    r'🏨']:
        m = re.search(pattern, html_content)
        if m and tourism_pos == -1:
            tourism_pos = m.start()

    for pattern in [r'Restaurant.*?(?:10 businesses|10 Businesses|Prospects)',
                    r'🍽']:
        m = re.search(pattern, html_content)
        if m and restaurant_pos == -1:
            restaurant_pos = m.start()

    for pattern in [r'Health.*?(?:10 businesses|10 Businesses|Prospects)',
                    r'🏥']:
        m = re.search(pattern, html_content)
        if m and health_pos == -1:
            health_pos = m.start()

    for pattern in [r'Beauty.*?(?:10 businesses|10 Businesses|Prospects)',
                    r'💇']:
        m = re.search(pattern, html_content)
        if m and beauty_pos == -1:
            beauty_pos = m.start()

    # Now extract each card
    for i, card_html in enumerate(cards):
        biz = {}
        biz['city'] = CITY_DISPLAY.get(city_slug, city_slug)

        # Determine vertical based on card position in original HTML
        card_start_snippet = card_html[:100].strip()
        card_pos = html_content.find(card_start_snippet)
        if card_pos == -1:
            card_pos = i * 1000

        positions = [
            (tourism_pos, "Tourism"),
            (restaurant_pos, "Restaurants"),
            (health_pos, "Health"),
            (beauty_pos, "Beauty")
        ]
        positions = [(p, v) for p, v in positions if p >= 0]
        positions.sort(key=lambda x: x[0])

        vertical = "Unknown"
        for pos, v in reversed(positions):
            if card_pos > pos:
                vertical = v
                break

        if vertical == "Unknown":
            if i < 10: vertical = "Tourism"
            elif i < 20: vertical = "Restaurants"
            elif i < 30: vertical = "Health"
            else: vertical = "Beauty"

        biz['vertical'] = vertical

        # Extract business name from h3
        name_match = re.search(r'<h3[^>]*>(.*?)</h3>', card_html)
        biz['name'] = clean_html(name_match.group(1)) if name_match else f"Business {i+1}"
        # Clean numbered prefix like "1. " or "2. "
        biz['name'] = re.sub(r'^\d+\.\s*', '', biz['name'])

        # ===== HANDLE BOTH HTML FORMATS =====
        if card_format == 'prospect-card':
            # Prospect-card format: uses meta-row with label/value pairs
            # Extract URL from <div class="url"> or <p class="url"> or meta-row
            url_match = re.search(r'(?:<div|<p) class="url[^"]*">(.*?)(?:</div>|</p>)', card_html, re.DOTALL)
            if url_match:
                biz['website'] = clean_html(url_match.group(1))
            else:
                # Try meta-row Website
                web_match = re.search(r'Website.*?<a[^>]*href="([^"]*)"', card_html, re.DOTALL)
                if web_match:
                    url = web_match.group(1)
                    biz['website'] = url.replace('https://', '').replace('http://', '').replace('www.', '').rstrip('/')
                else:
                    web_match2 = re.search(r'Website.*?>([\w\.\-]+\.\w+)', card_html, re.DOTALL)
                    biz['website'] = clean_html(web_match2.group(1)) if web_match2 else "No website"

            # Extract email from meta-row
            email_matches = re.findall(r'Email[^<]*</span><span class="value[^"]*">(.*?)</span>', card_html, re.DOTALL)
            if not email_matches:
                email_matches = re.findall(r'mailto:([^"]+)"', card_html)
            if not email_matches:
                email_matches = re.findall(r'Email.*?>([\w\.\-\+]+@[\w\.\-]+)', card_html, re.DOTALL)

            if email_matches:
                email_text = clean_html(email_matches[0])
                if 'none' in email_text.lower() or 'not found' in email_text.lower() or 'no email' in email_text.lower():
                    biz['email'] = 'Not found'
                else:
                    # Extract actual email
                    em = re.search(r'[\w\.\-\+]+@[\w\.\-]+', email_text)
                    biz['email'] = em.group(0) if em else email_text
            else:
                biz['email'] = 'Not found'

            # Extract phone from meta-row
            phone_match = re.search(r'Phone.*?>([\+\d\(\)\s\-/]+)', card_html, re.DOTALL)
            biz['phone'] = clean_html(phone_match.group(1)).strip() if phone_match else ""

            # Extract address
            addr_match = re.search(r'Address.*?>(.*?)(?:</div>|</span>)', card_html, re.DOTALL)
            biz['address'] = clean_html(addr_match.group(1)) if addr_match else ""

            # Extract platform/tech
            platform_match = re.search(r'(?:Platform|Tech|Tech Stack).*?>(.*?)(?:</div>|</span>)', card_html, re.DOTALL)
            biz['tech_stack'] = clean_html(platform_match.group(1)) if platform_match else ""

            # Extract analytics
            analytics_match = re.search(r'Analytics.*?>(.*?)(?:</div>|</span>)', card_html, re.DOTALL)
            biz['analytics_raw'] = clean_html(analytics_match.group(1)) if analytics_match else ""

            # Extract booking
            booking_match = re.search(r'Booking.*?>(.*?)(?:</div>|</span>)', card_html, re.DOTALL)
            biz['booking_raw'] = clean_html(booking_match.group(1)) if booking_match else ""

            # Extract social from meta-row
            social_match = re.search(r'Social.*?>(.*?)(?:</div>|</span>)', card_html, re.DOTALL)
            biz['social_raw'] = clean_html(social_match.group(1)) if social_match else ""

            # Extract CRM
            crm_match = re.search(r'CRM.*?>(.*?)(?:</div>|</span>)', card_html, re.DOTALL)
            biz['crm_raw'] = clean_html(crm_match.group(1)) if crm_match else ""

            # Extract description from email-section / Cold Email / NWM Opportunity
            desc_match = re.search(r'(?:Cold Email Angle|NWM Opportunity).*?<p>(.*?)</p>', card_html, re.DOTALL)
            biz['description'] = clean_html(desc_match.group(1))[:200] if desc_match else ""

            # Extract gap tags
            gap_tags = re.findall(r'class="gap-tag[^"]*">(.*?)</span>', card_html)
            gap_tags += re.findall(r'class="tag[^"]*">(.*?)</span>', card_html)
            biz['gaps'] = ' | '.join([clean_html(g) for g in gap_tags])

            # Extract tags
            tags = re.findall(r'class="tag[^"]*">(.*?)</span>', card_html)
            biz['tags'] = ' | '.join([clean_html(t) for t in tags])

            # Pitch
            biz['pitch'] = biz['description']  # Use the cold email angle as pitch
            biz['contact_raw'] = f"{biz['email']} · {biz['phone']}"

        else:
            # ===== ORIGINAL biz-card FORMAT =====
            # Try multiple URL extraction methods
            url_match = re.search(r'<div class="url[^"]*">(.*?)</div>', card_html, re.DOTALL)
            if url_match:
                biz['website'] = clean_html(url_match.group(1))
            else:
                # Try multiple Website extraction patterns for various HTML formats
                # Pattern A: contact-row / row with Website label + <a href>
                web_match = re.search(r'Website.*?<a[^>]*href="([^"]*)"[^>]*>(.*?)</a>', card_html, re.DOTALL)
                if web_match:
                    biz['website'] = clean_html(web_match.group(2))
                    if not biz['website'] or biz['website'] == '':
                        biz['website'] = web_match.group(1).replace('https://', '').replace('http://', '').replace('www.', '').rstrip('/')
                else:
                    # Pattern B: Website label followed by plain text domain
                    web_match2 = re.search(r'Website.*?>([\w\.\-]+\.[\w]{2,})', card_html, re.DOTALL)
                    if web_match2:
                        biz['website'] = clean_html(web_match2.group(1))
                    else:
                        # Pattern C: <p class="url"> format
                        web_match3 = re.search(r'<p class="url[^"]*">(.*?)</p>', card_html, re.DOTALL)
                        if web_match3:
                            biz['website'] = clean_html(web_match3.group(1))
                        else:
                            # Pattern D: <strong>Web:</strong> text (Valdivia format)
                            web_match4 = re.search(r'<strong>Web:</strong>\s*(.*?)(?:</div>|<)', card_html, re.DOTALL)
                            if web_match4:
                                biz['website'] = clean_html(web_match4.group(1))
                            else:
                                # Check if "No website" indicated
                                if re.search(r'(?:No website|NO WEBSITE|Has website.*?No|NO WEB)', card_html, re.IGNORECASE):
                                    biz['website'] = "No website"
                                else:
                                    biz['website'] = "No website"

        # Determine website status
        url_raw = biz['website'].upper()
        if 'NO WEBSITE' in url_raw or 'NONE' in url_raw:
            biz['website_status'] = 'No Website'
        elif 'DEAD' in url_raw or 'ECONNREFUSED' in url_raw or 'SOLD' in url_raw or 'FOR SALE' in url_raw:
            biz['website_status'] = 'Dead/Broken'
        elif 'SSL' in url_raw and ('ERROR' in url_raw or 'SELF-SIGNED' in url_raw or 'CERTIFICATE' in url_raw):
            biz['website_status'] = 'SSL Error'
        elif 'WIXSITE.COM' in url_raw or 'FREE SUBDOMAIN' in url_raw:
            biz['website_status'] = 'Free Subdomain'
        elif 'REDIRECT' in url_raw or '301' in url_raw:
            biz['website_status'] = 'Redirects'
        elif 'WIX' in url_raw.lower() or 'wix' in biz['website'].lower():
            biz['website_status'] = 'Active (Wix)'
        elif any(x in url_raw for x in ['SETMORE', 'AGENDAPRO', 'BOOKING PAGE ONLY']):
            biz['website_status'] = 'Platform Page Only'
        else:
            biz['website_status'] = 'Active'

        # Extract tags for tech/platform info
        tags = re.findall(r'<span class="tag[^"]*">(.*?)</span>', card_html)
        tag_texts = [clean_html(t) for t in tags]
        biz['tags'] = ' | '.join(tag_texts)

        # Extract tech details
        tech_match = re.search(r'<strong>Tech:</strong>(.*?)(?:</p>|<strong>)', card_html, re.DOTALL)
        biz['tech_stack'] = clean_html(tech_match.group(1)) if tech_match else ""

        # Extract gaps
        gaps_match = re.search(r'<strong>Gaps:</strong>(.*?)(?:</p>|<strong>)', card_html, re.DOTALL)
        biz['gaps'] = clean_html(gaps_match.group(1)) if gaps_match else ""

        # Extract address (try multiple patterns)
        addr_match = re.search(r'<strong>Address:</strong>(.*?)(?:</p>|<strong>)', card_html, re.DOTALL)
        if not addr_match:
            addr_match = re.search(r'Address:.*?</span>(.*?)(?:</div>|</a>)', card_html, re.DOTALL)
        biz['address'] = clean_html(addr_match.group(1)) if addr_match else ""

        # Extract reviews/description
        rev_match = re.search(r'<strong>Reviews:</strong>(.*?)(?:</p>|<strong>)', card_html, re.DOTALL)
        if not rev_match:
            # Try Digital Gaps or loss-box for description
            rev_match = re.search(r'Digital Gaps Found.*?</div>(.*?)(?:<div class="loss|<div class="audit)', card_html, re.DOTALL)
        biz['description'] = clean_html(rev_match.group(1))[:200] if rev_match else ""

        # Extract contact info
        contact_match = re.search(r'<div class="biz-contact">(.*?)</div>', card_html, re.DOTALL)
        contact_text = clean_html(contact_match.group(1)) if contact_match else ""
        biz['contact_raw'] = contact_text

        # Extract email — try multiple patterns
        email = None
        # Pattern 1: <span class="email">
        email_match = re.search(r'<span class="email[^"]*">(.*?)</span>', card_html)
        if email_match:
            email_text = clean_html(email_match.group(1))
            if 'no email' not in email_text.lower() and 'none' not in email_text.lower():
                email = email_text
        # Pattern 2: mailto:
        if not email:
            mailto_match = re.search(r'mailto:([\w\.\-\+]+@[\w\.\-]+)', card_html)
            if mailto_match:
                email = mailto_match.group(1)
        # Pattern 3: Email label + value span (row format) or <strong>Email:</strong> (detail format)
        if not email:
            em_match = re.search(r'<strong>Email:</strong>\s*(.*?)(?:</div>|<)', card_html, re.DOTALL)
            if em_match:
                email_text = clean_html(em_match.group(1))
                em_addr = re.search(r'[\w\.\-\+]+@[\w\.\-]+', email_text)
                if em_addr:
                    email = em_addr.group(0)
        if not email:
            em_match = re.search(r'Email.*?<span class="value">(.*?)</span>', card_html, re.DOTALL)
            if em_match:
                email_text = clean_html(em_match.group(1))
                em_addr = re.search(r'[\w\.\-\+]+@[\w\.\-]+', email_text)
                if em_addr:
                    email = em_addr.group(0)
                elif 'no email' not in email_text.lower() and 'none' not in email_text.lower() and 'not found' not in email_text.lower():
                    email = email_text
        # Pattern 4: Email: contact-row
        if not email:
            em_match = re.search(r'Email:.*?([\w\.\-\+]+@[\w\.\-]+)', card_html, re.DOTALL)
            if em_match:
                email = em_match.group(1)
        # Pattern 5: any email in card
        if not email:
            em_match = re.search(r'[\w\.\-\+]+@[\w\.\-]+\.\w{2,}', card_html)
            if em_match:
                email = em_match.group(0)

        biz['email'] = email if email else 'Not found'

        # Extract phone — try multiple patterns
        phone = None
        # Pattern 1: from contact div
        if contact_text:
            phone_match = re.search(r'(?:·|•|\|)\s*([\+\(\d][\d\s\(\)\-\+/]+)', contact_text)
            if phone_match:
                phone = phone_match.group(1).strip()
        # Pattern 2a: <strong>Phone:</strong> detail format
        if not phone:
            phone_match = re.search(r'<strong>Phone:</strong>\s*(.*?)(?:</div>|<)', card_html, re.DOTALL)
            if phone_match:
                phone = clean_html(phone_match.group(1)).strip()
        # Pattern 2b: Phone label + value span (row format)
        if not phone:
            phone_match = re.search(r'Phone.*?<span class="value">(.*?)</span>', card_html, re.DOTALL)
            if phone_match:
                phone = clean_html(phone_match.group(1)).strip()
        # Pattern 3: Phone: row
        if not phone:
            phone_match = re.search(r'Phone:.*?</span>([\+\d\(\)\s\-/]+)', card_html, re.DOTALL)
            if phone_match:
                phone = clean_html(phone_match.group(1)).strip()
        # Pattern 4: Chilean phone pattern
        if not phone:
            phone_match = re.search(r'(\+?56[\s\d\(\)\-]+|\(\d{2,3}\)\s*\d[\d\s\-]+)', card_html)
            if phone_match:
                phone = phone_match.group(1).strip()
        biz['phone'] = phone if phone else ""

        # Extract tech stack (try multiple patterns)
        if not biz.get('tech_stack'):
            tech_match = re.search(r'<strong>Tech:</strong>(.*?)(?:</p>|<strong>)', card_html, re.DOTALL)
            if tech_match:
                biz['tech_stack'] = clean_html(tech_match.group(1))
            else:
                # Try detail format: <strong>Tech:</strong>
                tech_match_detail = re.search(r'<strong>Tech:</strong>\s*(.*?)(?:</div>|<)', card_html, re.DOTALL)
                if tech_match_detail:
                    biz['tech_stack'] = clean_html(tech_match_detail.group(1))
                # Try row format: Tech label + value
                if not biz.get('tech_stack'):
                    tech_match2 = re.search(r'Tech.*?<span class="value">(.*?)</span>', card_html, re.DOTALL)
                if tech_match2:
                    biz['tech_stack'] = clean_html(tech_match2.group(1))
                else:
                    # Try audit-row format
                    audit_rows = re.findall(r'<span class="metric">(.*?)</span>.*?<span class="val[^"]*">(.*?)</span>', card_html, re.DOTALL)
                    if audit_rows:
                        biz['tech_stack'] = ' | '.join([f"{clean_html(m)}: {clean_html(v)}" for m, v in audit_rows])
                    else:
                        biz['tech_stack'] = ""

        # Extract gaps (try multiple patterns)
        if not biz.get('gaps'):
            gaps_match = re.search(r'<strong>Gaps:</strong>(.*?)(?:</p>|<strong>)', card_html, re.DOTALL)
            if gaps_match:
                biz['gaps'] = clean_html(gaps_match.group(1))
            else:
                gap_tags = re.findall(r'class="gap-tag[^"]*">(.*?)</span>', card_html)
                if gap_tags:
                    biz['gaps'] = ' | '.join([clean_html(g) for g in gap_tags])
                else:
                    biz['gaps'] = ""

        # Extract tags if not set
        if not biz.get('tags'):
            tags = re.findall(r'class="(?:tag|rec-tag|gap-tag)[^"]*">(.*?)</span>', card_html)
            biz['tags'] = ' | '.join([clean_html(t) for t in tags])

        # Extract pitch
        if not biz.get('pitch'):
            pitch_match = re.search(r'<strong>Pitch:</strong>(.*?)(?:</div>|$)', card_html, re.DOTALL)
            if not pitch_match:
                pitch_match = re.search(r'Cold Email Angle.*?<p>(.*?)</p>', card_html, re.DOTALL)
            if not pitch_match:
                pitch_match = re.search(r'NWM Opportunity.*?<p>(.*?)</p>', card_html, re.DOTALL)
            biz['pitch'] = clean_html(pitch_match.group(1))[:300] if pitch_match else ""

        # Determine analytics status
        tech_upper = biz['tech_stack'].upper()
        if 'GA4' in tech_upper or 'G-' in tech_upper:
            if 'DEAD' in tech_upper:
                biz['analytics'] = 'Dead GA4'
            else:
                biz['analytics'] = 'GA4 Active'
        elif 'UA-' in tech_upper:
            if 'DEAD' in tech_upper:
                biz['analytics'] = 'Dead UA (since 2023)'
            else:
                biz['analytics'] = 'UA (deprecated)'
        elif 'ZERO ANALYTICS' in tech_upper or 'NO ANALYTICS' in tech_upper or 'ZERO' in biz['gaps'].upper():
            biz['analytics'] = 'None'
        elif biz['website_status'] in ['No Website', 'Dead/Broken', 'SSL Error']:
            biz['analytics'] = 'N/A'
        else:
            biz['analytics'] = 'None'

        # Determine social media presence
        social_indicators = []
        combined = (biz['tech_stack'] + ' ' + biz['gaps'] + ' ' + card_html).lower()
        if 'facebook' in combined or 'fb pixel' in combined or 'fb @' in combined or 'fb sdk' in combined:
            social_indicators.append('Facebook')
        if 'instagram' in combined or 'ig @' in combined or '@' in combined and 'ig' in combined:
            social_indicators.append('Instagram')
        if 'youtube' in combined:
            social_indicators.append('YouTube')
        if 'tiktok' in combined:
            social_indicators.append('TikTok')

        biz['social_media'] = ', '.join(social_indicators) if social_indicators else 'None found'

        # Determine FB Pixel
        if 'fb pixel' in combined and ('no fb pixel' not in combined):
            biz['fb_pixel'] = 'Yes'
        else:
            biz['fb_pixel'] = 'No'

        # Determine CMS/Platform
        tech_lower = biz['tech_stack'].lower()
        if 'wordpress' in tech_lower or 'wp ' in tech_lower or 'wp+' in tech_lower:
            biz['platform'] = 'WordPress'
        elif 'wix' in tech_lower:
            biz['platform'] = 'Wix'
        elif 'squarespace' in tech_lower:
            biz['platform'] = 'Squarespace'
        elif 'shopify' in tech_lower:
            biz['platform'] = 'Shopify'
        elif 'joomla' in tech_lower:
            biz['platform'] = 'Joomla'
        elif 'static' in tech_lower or 'html5' in tech_lower:
            biz['platform'] = 'Static HTML'
        elif 'php' in tech_lower:
            biz['platform'] = 'Custom PHP'
        elif biz['website_status'] == 'No Website':
            biz['platform'] = 'N/A'
        elif biz['website_status'] in ['Dead/Broken', 'SSL Error']:
            biz['platform'] = 'Unknown (Dead)'
        else:
            biz['platform'] = 'Unknown'

        # Determine booking system
        booking_keywords = ['booking', 'reserv', 'agendapro', 'fresha', 'octopus24', 'woki',
                          'coverManager', 'setmore', 'dentalink', 'dentidesk', 'healthatom',
                          'mphb', 'direct-book', 'flow.cl', 'webpay']
        has_booking = any(kw in combined for kw in booking_keywords)
        if has_booking:
            # Find which booking system
            booking_systems = []
            for kw in ['AgendaPro', 'Fresha', 'Octopus24', 'Woki', 'CoverManager', 'Setmore',
                       'Dentalink', 'Dentidesk', 'HealthAtom', 'MPHB', 'Direct-book', 'Flow.cl',
                       'Webpay', 'Booking.com', 'TravelClick']:
                if kw.lower() in combined:
                    booking_systems.append(kw)
            biz['booking'] = ', '.join(booking_systems) if booking_systems else 'Yes (unspecified)'
        else:
            biz['booking'] = 'None'

        # Has CRM?
        crm_keywords = ['crm', 'hubspot', 'salesforce', 'bitrix', 'brevo', 'mailchimp', 'gohighlevel']
        has_crm = any(kw in combined for kw in crm_keywords)
        if has_crm and 'no crm' not in combined:
            biz['crm'] = 'Yes'
        else:
            biz['crm'] = 'No'

        # Extract pitch
        pitch_match = re.search(r'<strong>Pitch:</strong>(.*?)(?:</div>|$)', card_html, re.DOTALL)
        biz['pitch'] = clean_html(pitch_match.group(1)) if pitch_match else ""

        # Has Schema.org?
        if 'schema.org' in combined or 'schema' in combined:
            if 'no schema' in combined:
                biz['schema'] = 'No'
            else:
                biz['schema'] = 'Yes'
        else:
            biz['schema'] = 'No'

        # For prospect-card format, derive standard fields from raw data
        if card_format == 'prospect-card':
            # Website status
            url_raw = biz['website'].upper()
            if 'NO WEBSITE' in url_raw or biz['website'] == 'No website' or 'NONE' in url_raw:
                biz['website_status'] = 'No Website'
            elif 'DEAD' in url_raw or 'ECONNREFUSED' in url_raw or 'SOLD' in url_raw or 'FOR SALE' in url_raw:
                biz['website_status'] = 'Dead/Broken'
            elif 'SSL' in url_raw and ('ERROR' in url_raw or 'CERTIFICATE' in url_raw):
                biz['website_status'] = 'SSL Error'
            elif '403' in url_raw or 'FORBIDDEN' in url_raw:
                biz['website_status'] = 'Dead/Broken'
            elif 'WIXSITE.COM' in url_raw or 'FREE SUBDOMAIN' in url_raw:
                biz['website_status'] = 'Free Subdomain'
            elif 'REDIRECT' in url_raw or '301' in url_raw:
                biz['website_status'] = 'Redirects'
            elif 'SETMORE' in url_raw or 'AGENDAPRO' in url_raw:
                biz['website_status'] = 'Platform Page Only'
            elif biz['website'] and biz['website'] != 'No website':
                biz['website_status'] = 'Active'
            else:
                biz['website_status'] = 'No Website'

            # Analytics
            ar = biz.get('analytics_raw', '').upper()
            if 'GA4' in ar or 'G-' in ar:
                biz['analytics'] = 'GA4 Active'
            elif 'GTM' in ar and 'GA' in ar:
                biz['analytics'] = 'GA4 Active'
            elif 'NONE' in ar or 'ZERO' in ar or 'NO ' in ar:
                biz['analytics'] = 'None'
            elif 'DEAD' in ar:
                biz['analytics'] = 'Dead UA (since 2023)'
            elif ar:
                biz['analytics'] = ar[:30]
            else:
                biz['analytics'] = 'None' if biz['website_status'] not in ['No Website', 'Dead/Broken'] else 'N/A'

            # Platform
            ts = biz['tech_stack'].lower()
            if 'wordpress' in ts or 'wp' in ts or 'elementor' in ts:
                biz['platform'] = 'WordPress'
            elif 'wix' in ts:
                biz['platform'] = 'Wix'
            elif 'squarespace' in ts:
                biz['platform'] = 'Squarespace'
            elif 'shopify' in ts:
                biz['platform'] = 'Shopify'
            elif 'static' in ts or 'html' in ts:
                biz['platform'] = 'Static HTML'
            elif biz['website_status'] == 'No Website':
                biz['platform'] = 'N/A'
            elif biz['website_status'] in ['Dead/Broken', 'SSL Error']:
                biz['platform'] = 'Unknown (Dead)'
            elif ts:
                biz['platform'] = ts[:25]
            else:
                biz['platform'] = 'Unknown'

            # Social media
            sr = biz.get('social_raw', '').lower()
            combined = (card_html + sr).lower()
            social_list = []
            if 'facebook' in combined or 'fb' in combined:
                social_list.append('Facebook')
            if 'instagram' in combined or 'ig' in combined:
                social_list.append('Instagram')
            if 'youtube' in combined:
                social_list.append('YouTube')
            if 'tiktok' in combined:
                social_list.append('TikTok')
            biz['social_media'] = ', '.join(social_list) if social_list else 'None found'

            # FB Pixel
            if 'fb pixel' in combined or 'facebook pixel' in combined:
                if 'no fb pixel' in combined or 'no pixel' in combined:
                    biz['fb_pixel'] = 'No'
                else:
                    biz['fb_pixel'] = 'Yes'
            else:
                biz['fb_pixel'] = 'No'

            # Booking
            br = biz.get('booking_raw', '').lower()
            if 'none' in br or 'no ' in br or not br:
                biz['booking'] = 'None'
            else:
                biz['booking'] = biz.get('booking_raw', 'None')

            # CRM
            cr = biz.get('crm_raw', '').lower()
            if 'none' in cr or 'no' in cr or not cr:
                biz['crm'] = 'No'
            else:
                biz['crm'] = 'Yes'

            # Schema
            if 'schema' in combined:
                if 'no schema' in combined:
                    biz['schema'] = 'No'
                else:
                    biz['schema'] = 'Yes'
            else:
                biz['schema'] = 'No'

        # Skip template/pricing/non-business cards
        skip_keywords = ['template', 'starter', 'growth', 'fractional cmo', '$497', '$997', '$1,997', 'pricing']
        name_lower = biz['name'].lower()
        if any(kw in name_lower for kw in skip_keywords):
            continue

        businesses.append(biz)

    # Limit to 40 per city (in case of extra cards)
    return businesses[:40]

def create_xlsx(all_businesses):
    """Create the XLSX workbook."""
    wb = Workbook()

    # ===== SHEET 1: All Businesses Master List =====
    ws = wb.active
    ws.title = "All 680 Businesses"

    headers = [
        "City", "Vertical", "Business Name", "Website URL", "Website Status",
        "Platform/CMS", "Email", "Phone", "Address",
        "Analytics", "FB Pixel", "Social Media", "Booking System",
        "CRM", "Schema.org", "Tech Stack Summary", "Digital Gaps",
        "Description", "Sales Pitch"
    ]

    # Styles
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0D1117', end_color='0D1117', fill_type='solid')
    brand_fill = PatternFill(start_color='FF6B00', end_color='FF6B00', fill_type='solid')
    red_font = Font(name='Arial', color='F85149', size=10)
    green_font = Font(name='Arial', color='3FB950', size=10)
    normal_font = Font(name='Arial', size=10)
    yellow_font = Font(name='Arial', color='D29922', size=10)
    thin_border = Border(
        left=Side(style='thin', color='30363D'),
        right=Side(style='thin', color='30363D'),
        top=Side(style='thin', color='30363D'),
        bottom=Side(style='thin', color='30363D')
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = brand_fill if col <= 3 else header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Write data
    for row_idx, biz in enumerate(all_businesses, 2):
        data = [
            biz['city'], biz['vertical'], biz['name'], biz['website'], biz['website_status'],
            biz['platform'], biz['email'], biz['phone'], biz['address'],
            biz['analytics'], biz['fb_pixel'], biz['social_media'], biz['booking'],
            biz['crm'], biz['schema'], biz['tech_stack'], biz['gaps'],
            biz['description'], biz['pitch']
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

            # Color-code website status
            if col == 5:  # Website Status
                if value in ['Dead/Broken', 'SSL Error']:
                    cell.font = red_font
                elif value == 'No Website':
                    cell.font = red_font
                elif value == 'Active':
                    cell.font = green_font
                elif value in ['Free Subdomain', 'Redirects', 'Platform Page Only']:
                    cell.font = yellow_font

            # Color-code analytics
            if col == 10:  # Analytics
                if value in ['None', 'N/A', 'Dead UA (since 2023)', 'Dead GA4']:
                    cell.font = red_font
                elif value == 'GA4 Active':
                    cell.font = green_font

            # Color-code CRM
            if col == 14:  # CRM
                if value == 'No':
                    cell.font = red_font
                elif value == 'Yes':
                    cell.font = green_font

            # Color-code FB Pixel
            if col == 11:  # FB Pixel
                if value == 'No':
                    cell.font = red_font
                elif value == 'Yes':
                    cell.font = green_font

    # Column widths
    col_widths = [15, 12, 30, 35, 15, 15, 30, 20, 35, 18, 10, 25, 20, 8, 10, 50, 50, 40, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(all_businesses)+1}"

    # ===== SHEET 2: Summary by City =====
    ws2 = wb.create_sheet("City Summary")

    city_headers = [
        "City", "Total Businesses", "No Website", "Dead/SSL", "Active Sites",
        "Gmail/No Email", "Has Analytics", "Has FB Pixel", "Has Booking",
        "Has CRM", "Est. Revenue Lost"
    ]

    for col, header in enumerate(city_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = brand_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    ws2.freeze_panes = 'A2'

    revenue_map = {
        "Santiago": "$2.4M", "Valparaíso": "$2.1M", "Concepción": "$1.9M",
        "Antofagasta": "$2.2M", "Iquique": "$2.1M", "La Serena": "$1.82M",
        "Rancagua": "$1.95M", "Temuco": "$2.0M", "Puerto Montt": "$2.05M",
        "Arica": "$2.15M", "Copiapó": "$2.1M", "Talca": "$1.95M",
        "Chillán": "$2.0M", "Osorno": "$2.1M", "Valdivia": "$2.05M",
        "Punta Arenas": "$2.2M", "Coyhaique": "$2.05M"
    }

    city_order = list(CITY_DISPLAY.values())
    row = 2
    for city_name in city_order:
        city_biz = [b for b in all_businesses if b['city'] == city_name]
        if not city_biz:
            continue

        no_website = sum(1 for b in city_biz if b['website_status'] == 'No Website')
        dead_ssl = sum(1 for b in city_biz if b['website_status'] in ['Dead/Broken', 'SSL Error'])
        active = sum(1 for b in city_biz if b['website_status'] in ['Active', 'Active (Wix)'])
        gmail_no_email = sum(1 for b in city_biz if 'gmail' in b['email'].lower() or b['email'] == 'Not found')
        has_analytics = sum(1 for b in city_biz if 'Active' in b['analytics'])
        has_fb = sum(1 for b in city_biz if b['fb_pixel'] == 'Yes')
        has_booking = sum(1 for b in city_biz if b['booking'] != 'None')
        has_crm = sum(1 for b in city_biz if b['crm'] == 'Yes')

        data = [
            city_name, len(city_biz), no_website, dead_ssl, active,
            gmail_no_email, has_analytics, has_fb, has_booking,
            has_crm, revenue_map.get(city_name, "N/A")
        ]

        for col, value in enumerate(data, 1):
            cell = ws2.cell(row=row, column=col, value=value)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        row += 1

    # Totals row
    total_row = row
    ws2.cell(row=total_row, column=1, value="TOTAL (17 Cities)").font = Font(name='Arial', bold=True, size=11)
    ws2.cell(row=total_row, column=2, value=len(all_businesses)).font = Font(name='Arial', bold=True, size=11)
    ws2.cell(row=total_row, column=11, value="$35.12M").font = Font(name='Arial', bold=True, color='F85149', size=11)

    for col in range(3, 11):
        formula = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{total_row-1})"
        ws2.cell(row=total_row, column=col, value=formula).font = Font(name='Arial', bold=True, size=11)

    col_widths2 = [18, 16, 14, 12, 14, 16, 14, 14, 14, 10, 16]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ===== SHEET 3: Website & Social Audit =====
    ws3 = wb.create_sheet("Website & Social Audit")

    audit_headers = [
        "City", "Vertical", "Business Name", "Website URL", "Website Status",
        "Platform/CMS", "Has GA4", "Has FB Pixel", "Has GTM",
        "Has Schema.org", "Social: Facebook", "Social: Instagram",
        "Social: YouTube", "Social: TikTok", "Booking Platform",
        "Email Type", "Key Website Issues", "Key Social Issues"
    ]

    for col, header in enumerate(audit_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color='1A1E2E', end_color='1A1E2E', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    ws3.freeze_panes = 'A2'

    for row_idx, biz in enumerate(all_businesses, 2):
        # Determine individual social platforms
        combined = (biz.get('tech_stack', '') + ' ' + biz.get('gaps', '') + ' ' + biz.get('contact_raw', '')).lower()
        has_fb = 'Yes' if 'facebook' in combined or 'fb' in combined.split() or 'fb @' in combined or 'fb pixel' in combined else 'No'
        has_ig = 'Yes' if 'instagram' in combined or 'ig @' in combined else 'No'
        has_yt = 'Yes' if 'youtube' in combined else 'No'
        has_tt = 'Yes' if 'tiktok' in combined else 'No'
        has_gtm = 'Yes' if 'gtm' in combined else 'No'

        # Email type
        email = biz['email']
        if email == 'Not found':
            email_type = 'Not Found'
        elif 'gmail' in email.lower():
            email_type = 'Gmail (Unprofessional)'
        elif 'hotmail' in email.lower() or 'outlook' in email.lower():
            email_type = 'Hotmail/Outlook (Personal)'
        else:
            email_type = 'Professional Domain'

        # Key website issues
        website_issues = []
        gaps = biz.get('gaps', '').lower()
        if biz['website_status'] == 'No Website':
            website_issues.append('NO WEBSITE')
        if biz['website_status'] == 'Dead/Broken':
            website_issues.append('SITE DEAD/BROKEN')
        if biz['website_status'] == 'SSL Error':
            website_issues.append('SSL CERTIFICATE ERROR')
        if 'zero analytics' in gaps or 'no analytics' in gaps:
            website_issues.append('Zero analytics')
        if 'dead ua' in gaps or 'dead analytics' in gaps:
            website_issues.append('Dead analytics (~3yr)')
        if 'no booking' in gaps or 'no reservation' in gaps:
            website_issues.append('No booking system')
        if 'no crm' in gaps:
            website_issues.append('No CRM')
        if 'outdated' in gaps:
            website_issues.append('Outdated plugins')
        if 'free subdomain' in gaps or 'wixsite.com' in biz.get('website', '').lower():
            website_issues.append('Free subdomain (no SEO)')
        if 'split' in gaps or '2 domains' in gaps:
            website_issues.append('Split domain SEO')

        # Key social issues
        social_issues = []
        if has_fb == 'No' and has_ig == 'No':
            social_issues.append('NO social media presence')
        elif has_fb == 'Yes' and has_ig == 'No':
            social_issues.append('Facebook only, no Instagram')
        if biz['fb_pixel'] == 'No' and biz['website_status'] == 'Active':
            social_issues.append('No FB Pixel on website')
        if 'empty' in combined and 'fb' in combined:
            social_issues.append('FB SDK empty/broken')
        if biz['website_status'] == 'No Website' and (has_fb == 'Yes' or has_ig == 'Yes'):
            social_issues.append('Social without website = no conversion')

        data = [
            biz['city'], biz['vertical'], biz['name'], biz['website'], biz['website_status'],
            biz['platform'],
            'Yes' if 'GA4 Active' in biz.get('analytics', '') else 'No',
            biz['fb_pixel'], has_gtm, biz['schema'],
            has_fb, has_ig, has_yt, has_tt,
            biz['booking'], email_type,
            ' | '.join(website_issues) if website_issues else 'Minor issues only',
            ' | '.join(social_issues) if social_issues else 'Adequate'
        ]

        for col, value in enumerate(data, 1):
            cell = ws3.cell(row=row_idx, column=col, value=value)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

            # Color code Yes/No
            if value == 'Yes':
                cell.font = green_font
            elif value == 'No' and col >= 7:
                cell.font = red_font

    col_widths3 = [15, 12, 30, 35, 15, 15, 10, 12, 10, 12, 12, 12, 12, 12, 20, 22, 45, 45]
    for i, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    ws3.auto_filter.ref = f"A1:{get_column_letter(len(audit_headers))}{len(all_businesses)+1}"

    return wb


def main():
    all_businesses = []

    for city_slug in CITIES:
        filepath = os.path.join(DEPLOY_DIR, f"{city_slug}-prospects-report.html")
        if not os.path.exists(filepath):
            print(f"WARNING: {filepath} not found, skipping {city_slug}")
            continue

        print(f"Parsing {city_slug}...")
        with open(filepath, 'r', encoding='utf-8') as f:
            html = f.read()

        businesses = extract_businesses_from_html(html, city_slug)
        print(f"  Found {len(businesses)} businesses")
        all_businesses.extend(businesses)

    print(f"\nTotal businesses extracted: {len(all_businesses)}")

    # Create XLSX
    wb = create_xlsx(all_businesses)

    output_path = os.path.join(DEPLOY_DIR, "NetWebMedia_Chile_680_Businesses_Website_Social_Audit.xlsx")
    wb.save(output_path)
    print(f"\nSaved to: {output_path}")

if __name__ == "__main__":
    main()
