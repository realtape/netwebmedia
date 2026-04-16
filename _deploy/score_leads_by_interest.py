#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Score 680 prospects by likelihood to be interested in NetWebMedia services.

Scoring factors:
- Has a website (has digital presence) → positive signal
- Website status is "Inactive" or "Error" → high pain, high interest
- No social media presence → digital gap = high interest
- Small business (SMB niche) → more price-sensitive but responsive

Usage:
    python score_leads_by_interest.py --top 10 --per-region
    # Output: top 10 leads per region (170 total across 17 regions)
"""
import json
import os
import argparse
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(HERE, "NetWebMedia_Chile_680_Businesses_Website_Social_Audit.xlsx")

def score_prospect(row_data):
    """Score a prospect on likelihood to buy NetWebMedia services (0-100)."""
    score = 0
    details = {}

    # Extract fields from XLSX row
    city = row_data.get("city", "").lower()
    vertical = row_data.get("vertical", "").lower()
    name = row_data.get("name", "")
    website_status = row_data.get("website_status", "").lower()
    platform = row_data.get("platform", "").lower()
    email = row_data.get("email")
    social_media = row_data.get("social_media", "").lower()
    analytics = row_data.get("analytics", "").lower()
    fb_pixel = row_data.get("fb_pixel", "").lower()

    # Factor 1: Website presence (basic digital readiness)
    if website_status:
        if website_status in ("error", "invalid", "broken", "offline", "no"):
            score += 40  # High pain: website is down
            details["website_status"] = "broken/error"
        elif website_status in ("inactive", "parked", "minimal"):
            score += 30  # Medium pain: website exists but unmaintained
            details["website_status"] = "inactive"
        elif website_status in ("active", "up"):
            score += 15  # Low signal: they have a website but maybe it's old
            details["website_status"] = "active"

    # Factor 2: Platform/CMS assessment
    if platform:
        platform_lower = platform.lower()
        if "wix" in platform_lower or "squarespace" in platform_lower:
            score += 20  # Cheap platform = limited budget but responsive
            details["platform"] = "budget-cms"
        elif "wordpress" in platform_lower or "drupal" in platform_lower:
            score += 10  # Standard open-source
            details["platform"] = "standard-cms"
        elif "custom" in platform_lower or "proprietário" in platform_lower or "unknown" in platform_lower:
            score += 5   # Hard to upgrade/improve
            details["platform"] = "custom/unknown"

    # Factor 3: Analytics presence (digital maturity indicator)
    if analytics:
        analytics_lower = analytics.lower()
        if "no" in analytics_lower or "none" in analytics_lower:
            score += 25  # Operating blind = high pain
            details["analytics"] = "missing"
        elif "google analytics 3" in analytics_lower or "ga3" in analytics_lower:
            score += 15  # Outdated version
            details["analytics"] = "outdated"
        elif "google analytics 4" in analytics_lower or "ga4" in analytics_lower:
            score += 5   # Modern but only first step
            details["analytics"] = "current"

    # Factor 4: Social media gaps (biggest NetWebMedia pain point)
    if social_media:
        social_lower = social_media.lower()
        platforms_present = sum(1 for p in ["facebook", "instagram", "tiktok", "youtube", "linkedin"] if p in social_lower)
        if platforms_present == 0:
            score += 35  # Complete digital invisibility = VERY high interest
            details["social_media"] = "none"
        elif platforms_present == 1:
            score += 20  # Only one platform = underexploiting
            details["social_media"] = "limited"
        elif platforms_present >= 3:
            score += 5   # Already multi-platform
            details["social_media"] = "multi-platform"

    # Factor 5: FB Pixel (retargeting capability)
    if fb_pixel:
        fb_lower = fb_pixel.lower()
        if "no" in fb_lower or "none" in fb_lower:
            score += 15  # Can't retarget = missing revenue
            details["fb_pixel"] = "missing"

    # Factor 6: Niche priority (SMB = top priority per user request)
    niche_key = row_data.get("niche_key", "")
    if niche_key == "smb":
        score += 20  # SMB is "more important" per user
        details["niche"] = "SMB"
    elif niche_key == "tourism":
        score += 10  # High seasonality, responsive to ads
    elif niche_key == "restaurants":
        score += 15  # High failure rate, responsive to online ordering
    elif niche_key in ("law_firms", "real_estate"):
        score += 5   # Lower response rates but high ticket

    # Factor 7: Email quality (direct outreach signal)
    if email:
        email_lower = email.lower()
        if "@gmail" in email_lower or "@hotmail" in email_lower or "@yahoo" in email_lower:
            score += 10  # Generic email = less professional, might need upgrade
            details["email"] = "generic"
        else:
            score += 15  # Business domain = professional signal
            details["email"] = "business_domain"

    # Normalize to 0-100
    score = min(100, max(0, score))
    return score, details

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--top", type=int, default=10, help="Top N leads per region")
    ap.add_argument("--per-region", action="store_true", help="Select top N per region")
    ap.add_argument("--per-niche-region", action="store_true", help="Select top N per niche per region")
    ap.add_argument("--output", default="top_leads.json", help="Output JSON file")
    args = ap.parse_args()

    print("Loading and scoring prospects...")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    all_scored = []
    for row_idx in range(2, ws.max_row + 1):
        if not ws[row_idx][2].value:  # Skip if no name
            continue

        row_data = {
            "city": str(ws[row_idx][0].value or "").lower().replace(" ", "-"),
            "vertical": str(ws[row_idx][1].value or ""),
            "name": str(ws[row_idx][2].value or ""),
            "website": str(ws[row_idx][3].value or ""),
            "website_status": str(ws[row_idx][4].value or ""),
            "platform": str(ws[row_idx][5].value or ""),
            "email": str(ws[row_idx][6].value or "") if ws[row_idx][6] else None,
            "phone": str(ws[row_idx][7].value or ""),
            "address": str(ws[row_idx][8].value or ""),
            "analytics": str(ws[row_idx][9].value or "") if ws[row_idx][9] else None,
            "fb_pixel": str(ws[row_idx][10].value or "") if ws[row_idx][10] else None,
            "social_media": str(ws[row_idx][11].value or "") if ws[row_idx][11] else None,
        }

        score, details = score_prospect(row_data)
        all_scored.append({
            "name": row_data["name"],
            "city": row_data["city"],
            "email": row_data["email"],
            "website": row_data["website"],
            "score": score,
            "details": details,
        })

    wb.close()

    # Filter based on mode
    if args.per_niche_region:
        # Group by niche + region, take top N from each
        # (requires loading crm_import.json for niche assignments)
        print("ERROR: --per-niche-region not yet implemented")
        return

    if args.per_region:
        # Group by region, take top N from each
        by_region = defaultdict(list)
        for scored in all_scored:
            by_region[scored["city"]].append(scored)

        selected = []
        for city in sorted(by_region.keys()):
            city_leads = sorted(by_region[city], key=lambda x: x["score"], reverse=True)[:args.top]
            selected.extend(city_leads)
            print(f"  {city:20s}: top {len(city_leads)} (scores {city_leads[0]['score']:.0f} — {city_leads[-1]['score']:.0f})")

        print(f"\nSelected {len(selected)} leads ({args.top} per region × {len(by_region)} regions)")
    else:
        # Global top N
        selected = sorted(all_scored, key=lambda x: x["score"], reverse=True)[:args.top]
        print(f"Selected top {len(selected)} leads (global)")

    # Write output
    print(f"\nWriting {args.output}...")
    with open(os.path.join(HERE, args.output), "w", encoding="utf-8") as f:
        json.dump(selected, f, indent=2, ensure_ascii=False)

    # Stats
    scores = [s["score"] for s in selected]
    print(f"Score distribution: avg={sum(scores)/len(scores):.0f}, min={min(scores):.0f}, max={max(scores):.0f}")

if __name__ == "__main__":
    main()
